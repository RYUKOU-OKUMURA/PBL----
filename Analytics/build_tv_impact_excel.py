#!/usr/bin/env python3
"""
テレビ放送前後の検索順位・アクセス比較と戦略概要を Excel に出力する。
"""
from __future__ import annotations

import os
import sys
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

try:
    from google.oauth2 import service_account
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import DateRange, Dimension, Metric, RunReportRequest
    from googleapiclient.discovery import build
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("pip install openpyxl google-analytics-data google-api-python-client google-auth python-dotenv", file=sys.stderr)
    raise SystemExit(1) from e

# テレビ放映日（データ上のスパイク起点）
TV_DAY = date(2026, 3, 18)

# 比較用キーワード（表記ゆれ含む）
FOCUS_QUERIES = [
    "レッドコード 整体",
    "レッドコード整体",
    "レッドコード施術",
    "レッドコード 施術",
    "レッドコード 名古屋",
    "レッド コード 整体",
    "無重力整体 レッドコード",
    "無重力整体レッドコード",
    "フィジカルバランスラボ整体院",
]


def gsc_service():
    creds_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    site = os.getenv("GSC_SITE_URL", "").strip()
    creds = service_account.Credentials.from_service_account_file(
        creds_path,
        scopes=["https://www.googleapis.com/auth/webmasters.readonly"],
    )
    return build("searchconsole", "v1", credentials=creds), site


def ga4_client():
    prop = os.getenv("GA4_PROPERTY_ID", "").strip()
    creds_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    creds = service_account.Credentials.from_service_account_file(creds_path)
    return BetaAnalyticsDataClient(credentials=creds), f"properties/{prop}"


def gsc_fetch_queries(start: date, end: date) -> dict[str, dict]:
    svc, site = gsc_service()
    body = {
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "dimensions": ["query"],
        "rowLimit": 25000,
    }
    rows = svc.searchanalytics().query(siteUrl=site, body=body).execute().get("rows", [])
    return {
        r["keys"][0]: {
            "clicks": r.get("clicks", 0),
            "impressions": r.get("impressions", 0),
            "position": r.get("position", 0),
            "ctr": r.get("ctr", 0),
        }
        for r in rows
    }


def ga4_sum_sessions_pv_users(start: date, end: date) -> dict:
    client, prop = ga4_client()
    req = RunReportRequest(
        property=prop,
        dimensions=[Dimension(name="date")],
        metrics=[
            Metric(name="sessions"),
            Metric(name="screenPageViews"),
            Metric(name="totalUsers"),
        ],
        date_ranges=[DateRange(start_date=start.isoformat(), end_date=end.isoformat())],
    )
    resp = client.run_report(req)
    sess = pv = users = 0
    for row in resp.rows or []:
        mv = row.metric_values
        sess += int(float(mv[0].value))
        pv += int(float(mv[1].value))
        users += int(float(mv[2].value))
    return {"sessions": sess, "screenPageViews": pv, "totalUsers": users, "days": (end - start).days + 1}


def style_header(ws, row: int, ncol: int):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, max_width: int = 45):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[letter].width = min(max(12, maxlen + 2), max_width)


def main() -> None:
    today = date.today()
    pre_end = TV_DAY - timedelta(days=1)
    post_start = TV_DAY

    # 公平比較：放送直前28日 vs 放送日〜本日
    pre_28_start = TV_DAY - timedelta(days=28)

    pre_map = gsc_fetch_queries(pre_28_start, pre_end)
    post_map = gsc_fetch_queries(post_start, today)

    # 追加：累計（放送前日まで）で「長期」の順位感
    pre_cum_map = gsc_fetch_queries(date(2024, 1, 1), pre_end)

    wb = Workbook()

    # --- Sheet 1: 検索順位・パフォーマンス ---
    ws1 = wb.active
    ws1.title = "検索キーワード比較"
    headers1 = [
        "キーワード",
        "放送前(28日)期間",
        "平均順位(前)",
        "クリック(前)",
        "インプレ(前)",
        "放送後期間",
        "平均順位(後)",
        "クリック(後)",
        "インプレ(後)",
        "順位差(後-前)※",
        "備考",
    ]
    for i, h in enumerate(headers1, 1):
        ws1.cell(1, i, h)
    style_header(ws1, 1, len(headers1))

    pre_period_str = f"{pre_28_start}〜{pre_end}"
    post_period_str = f"{post_start}〜{today}"

    row = 2
    for q in FOCUS_QUERIES:
        a = pre_map.get(q)
        b = post_map.get(q)
        pos_pre = round(a["position"], 2) if a else None
        pos_post = round(b["position"], 2) if b else None
        diff = (pos_post - pos_pre) if (a and b) else None
        diff_s = f"{diff:+.2f}" if diff is not None else "—"
        note = ""
        if not a and b:
            note = "放送前28日は検索ボリュームが少なく行未表示"
        elif a and not b:
            note = "放送後は期間が短い／表記ゆれ"
        ws1.cell(row, 1, q)
        ws1.cell(row, 2, pre_period_str)
        ws1.cell(row, 3, pos_pre if pos_pre is not None else "—")
        ws1.cell(row, 4, a["clicks"] if a else "—")
        ws1.cell(row, 5, a["impressions"] if a else "—")
        ws1.cell(row, 6, post_period_str)
        ws1.cell(row, 7, pos_post if pos_post is not None else "—")
        ws1.cell(row, 8, b["clicks"] if b else "—")
        ws1.cell(row, 9, b["impressions"] if b else "—")
        ws1.cell(row, 10, diff_s)
        ws1.cell(row, 11, note)
        row += 1

    # 補足行：累計比較（代表キーワードのみ）
    ws1.cell(row, 1, "【参考】累計(2024/1/1〜放送前日)")
    ws1.cell(row, 2, f"〜{pre_end}")
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    row += 1
    for q in ["レッドコード 整体", "レッドコード 名古屋"]:
        a = pre_cum_map.get(q)
        b = post_map.get(q)
        if a or b:
            ws1.cell(row, 1, q + " (累計vs放送後)")
            ws1.cell(row, 2, f"累計〜{pre_end}")
            ws1.cell(row, 3, round(a["position"], 2) if a else "—")
            ws1.cell(row, 4, a["clicks"] if a else "—")
            ws1.cell(row, 5, a["impressions"] if a else "—")
            ws1.cell(row, 6, post_period_str)
            ws1.cell(row, 7, round(b["position"], 2) if b else "—")
            ws1.cell(row, 8, b["clicks"] if b else "—")
            ws1.cell(row, 9, b["impressions"] if b else "—")
            d = (b["position"] - a["position"]) if (a and b) else None
            ws1.cell(row, 10, f"{d:+.2f}" if d is not None else "—")
            ws1.cell(row, 11, "GSCの平均掲載順位(加重)")
            row += 1

    ws1.cell(row, 1, "※順位差(後−前)：プラスは順位が下がった、マイナスは上がった(1位に近い)。GSCの平均掲載順位は日・端末で変動します。")
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    ws1.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")

    autosize_columns(ws1)

    # --- Sheet 2: アクセス数（GA4）---
    ws2 = wb.create_sheet("アクセス数の変化")
    ga_pre = ga4_sum_sessions_pv_users(pre_28_start, pre_end)
    ga_post = ga4_sum_sessions_pv_users(post_start, today)

    ws2.cell(1, 1, "指標")
    ws2.cell(1, 2, f"放送前28日\n{pre_28_start}〜{pre_end}")
    ws2.cell(1, 3, f"放送後\n{post_start}〜{today}")
    ws2.cell(1, 4, "1日あたり(前)")
    ws2.cell(1, 5, "1日あたり(後)")
    style_header(ws2, 1, 5)

    metrics = [
        ("セッション", "sessions"),
        ("ページビュー", "screenPageViews"),
        ("ユーザー数", "totalUsers"),
    ]
    for i, (label, key) in enumerate(metrics, 2):
        v_pre = ga_pre[key]
        v_post = ga_post[key]
        d_pre = ga_pre["days"]
        d_post = ga_post["days"]
        ws2.cell(i, 1, label)
        ws2.cell(i, 2, v_pre)
        ws2.cell(i, 3, v_post)
        ws2.cell(i, 4, round(v_pre / d_pre, 1))
        ws2.cell(i, 5, round(v_post / d_post, 1))

    note_row = 6
    ws2.cell(note_row, 1, "補足")
    ws2.cell(
        note_row + 1,
        1,
        "・放送後は日数が少ないため「合計」より「1日あたり」も参照してください。\n"
        "・テレビ特需が落ち着いた後の平常値は、別途同じ28日窓で再測定するとよいです。\n"
        "・チャネル別内訳は別レポート（Direct比率など）を参照。",
    )
    ws2.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 3, end_column=5)
    ws2.cell(note_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    autosize_columns(ws2, 50)

    # --- Sheet 3: 戦略概要 ---
    ws3 = wb.create_sheet("維持・強化の戦略概要")
    strategy = [
        ("目的", "認知拡大後も「検索での発見」と「サイトアクセス」を落とさず、来院・予約につなげる"),
        ("", ""),
        ("1. 検索順位（オーガニック）", ""),
        ("　・", "ブランド系：トップ／About／レッドコード説明ページのタイトル・ディスクリプションをクエリと整合。重複・食い違いを解消。"),
        ("　・", "症状・情報系：GSCで既にインプレがあるクエリから優先し、見出し・内部リンク・更新日で順位帯を上げる。"),
        ("　・", "地域ワード：オーガニックだけに頼らず、Googleビジネスプロフィール（MEO）と役割分担。ガイドの12カテゴリ投稿を継続。"),
        ("", ""),
        ("2. アクセスの維持", ""),
        ("　・", "計測：GA4で週次のセッション・チャネル・ランディングを定点観測。急落時はソース別に切り分け。"),
        ("　・", "コンテンツ：エビデンス記事・FAQを継続し、指名検索以外の入口も増やす。"),
        ("　・", "オフWeb：LINE・口コミ・紹介と連動し、検索単体に依存しない導線設計。"),
        ("", ""),
        ("3. リスク管理", ""),
        ("　・", "特需落ち：テレビ後の「新しい平常」を数値で把握し、目標を再設定する。"),
        ("　・", "順位変動：アルゴリズム・競合はコントロール不可のため、E-E-A-T（専門性・経験）と一次情報を積み上げる。"),
        ("", ""),
        ("4. 次のアクション例", ""),
        ("　・", "毎月：GSCのクエリTOP・平均順位・クリックの定点表を更新（本シートと同じ切り口）。"),
        ("　・", "四半期：症状系キーワードの網羅度レビュー、既存記事のリライト優先度付け。"),
    ]
    for r, (a, b) in enumerate(strategy, 1):
        ws3.cell(r, 1, a)
        ws3.cell(r, 2, b)
        if a and a[0].isdigit():
            ws3.cell(r, 1).font = Font(bold=True)
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 88
    for r in range(1, len(strategy) + 1):
        ws3.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")

    out_path = ROOT / "Analytics" / f"TV放映_検索順位とアクセス比較_{today.isoformat()}.xlsx"
    wb.save(out_path)
    print(f"保存: {out_path}")


if __name__ == "__main__":
    main()
